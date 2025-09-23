import ast
import csv
import json
import warnings
from os import listdir
from typing import Union, Literal, Iterable

import openpyxl as xl
from mango.logging import get_configured_logger
from openpyxl.utils import get_column_letter
from pytups import TupList

log = get_configured_logger(__name__)


def list_files_directory(directory: str, extensions: list = None):
    """
    List files in a directory with optional extension filtering.

    Returns a list of file paths from the specified directory, optionally
    filtered by file extensions. If no extensions are provided, all files
    in the directory are returned.

    :param directory: Directory path to search for files
    :type directory: str
    :param extensions: List of file extensions to filter by (e.g., ['.txt', '.csv'])
    :type extensions: list, optional
    :return: List of file paths matching the criteria
    :rtype: list[str]
    :raises OSError: If the directory doesn't exist or cannot be accessed

    Example:
        >>> list_files_directory('/path/to/files', ['.txt', '.csv'])
        ['/path/to/files/file1.txt', '/path/to/files/data.csv']
        >>> list_files_directory('/path/to/files')
        ['/path/to/files/file1.txt', '/path/to/files/data.csv', '/path/to/files/image.png']
    """
    if extensions is None:
        extensions = ["."]
    return [
        rf"{directory}\{f}"
        for f in listdir(rf"{directory}")
        if any([f.__contains__(f"{ext}") for ext in extensions])
    ]


def check_extension(path: str, extension: str):
    """
    Check if a file path has the specified extension.

    Performs a simple string check to determine if the file path
    ends with the specified extension.

    :param path: File path to check
    :type path: str
    :param extension: Extension to check for (e.g., '.txt', '.csv')
    :type extension: str
    :return: True if the file has the specified extension, False otherwise
    :rtype: bool

    Example:
        >>> check_extension('/path/to/file.txt', '.txt')
        True
        >>> check_extension('/path/to/file.csv', '.txt')
        False
    """
    return path.endswith(extension)


def is_excel_file(path: str):
    """
    Check if a file is an Excel file based on its extension.

    Determines if the file is an Excel file by checking if it has
    one of the common Excel file extensions (.xlsx, .xls, .xlsm).

    :param path: File path to check
    :type path: str
    :return: True if the file is an Excel file, False otherwise
    :rtype: bool

    Example:
        >>> is_excel_file('/path/to/data.xlsx')
        True
        >>> is_excel_file('/path/to/data.csv')
        False
    """
    return (
        check_extension(path, ".xlsx")
        or check_extension(path, ".xls")
        or check_extension(path, ".xlsm")
    )


def is_json_file(path: str):
    """
    Check if a file is a JSON file based on its extension.

    Determines if the file is a JSON file by checking if it has
    the .json extension.

    :param path: File path to check
    :type path: str
    :return: True if the file is a JSON file, False otherwise
    :rtype: bool

    Example:
        >>> is_json_file('/path/to/config.json')
        True
        >>> is_json_file('/path/to/data.csv')
        False
    """
    return check_extension(path, ".json")


def load_json(path: str, **kwargs):
    """
    Load a JSON file and return its contents as a Python object.

    Reads a JSON file from the specified path and parses it into
    a Python dictionary, list, or other JSON-compatible object.

    :param path: Path to the JSON file to load
    :type path: str
    :param kwargs: Additional keyword arguments passed to json.load()
    :return: Parsed JSON content (dict, list, etc.)
    :rtype: Union[dict, list, str, int, float, bool]
    :raises FileNotFoundError: If the file doesn't exist
    :raises json.JSONDecodeError: If the file contains invalid JSON

    Example:
        >>> data = load_json('/path/to/config.json')
        >>> print(data['setting'])
        'value'
    """
    with open(path, "r", **kwargs) as f:
        return json.load(f)


def write_json(data: Union[dict, list], path):
    """
    Write data to a JSON file with pretty formatting.

    Serializes a Python object (dict, list, etc.) to JSON format and
    writes it to the specified file with indentation for readability.

    :param data: Python object to serialize (dict, list, etc.)
    :type data: Union[dict, list]
    :param path: Path where the JSON file should be written
    :type path: str
    :return: None
    :raises TypeError: If the data cannot be serialized to JSON
    :raises OSError: If the file cannot be written

    Example:
        >>> data = {'name': 'John', 'age': 30}
        >>> write_json(data, '/path/to/output.json')
    """

    with open(path, "w") as f:
        json.dump(data, f, indent=4, sort_keys=False)


def load_excel_sheet(path: str, sheet: str, **kwargs):
    """
    Load a specific sheet from an Excel file as a pandas DataFrame.

    Reads a single sheet from an Excel file and returns it as a pandas
    DataFrame. Requires pandas to be installed.

    :param path: Path to the Excel file
    :type path: str
    :param sheet: Name of the sheet to load
    :type sheet: str
    :param kwargs: Additional keyword arguments passed to pandas.read_excel()
    :return: DataFrame containing the sheet data
    :rtype: pandas.DataFrame
    :raises FileNotFoundError: If the file is not an Excel file
    :raises NotImplementedError: If pandas is not installed
    :raises ValueError: If the specified sheet doesn't exist

    Example:
        >>> df = load_excel_sheet('/path/to/data.xlsx', 'Sheet1')
        >>> print(df.head())
    """
    try:
        import pandas as pd
    except ImportError as e:
        raise NotImplementedError("function not yet implemented without pandas")

    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    return pd.read_excel(path, sheet_name=sheet, **kwargs)


def load_excel(
    path,
    dtype="object",
    output: Literal[
        "df", "dict", "list", "series", "split", "tight", "records", "index"
    ] = "df",
    sheet_name=None,
    **kwargs,
):
    """
    Load an Excel file with flexible output format options.

    Reads an Excel file and returns the data in various formats. Can load
    all sheets or a specific sheet, and convert the output to different
    formats (DataFrame, dictionary, list of records, etc.).

    :param path: Path to the Excel file
    :type path: str
    :param dtype: Data type for columns (default: "object" to preserve original data)
    :type dtype: str or dict
    :param output: Output format ("df", "dict", "list", "records", etc.)
    :type output: Literal["df", "dict", "list", "series", "split", "tight", "records", "index"]
    :param sheet_name: Name of sheet to read (None for all sheets)
    :type sheet_name: str, optional
    :param kwargs: Additional arguments passed to pandas.read_excel()
    :return: Data in the specified output format
    :rtype: Union[pandas.DataFrame, dict, list]
    :raises FileNotFoundError: If the file is not an Excel file
    :raises ImportError: If pandas is not installed

    Example:
        >>> # Load all sheets as DataFrames
        >>> data = load_excel('/path/to/data.xlsx')
        >>>
        >>> # Load specific sheet as list of records
        >>> data = load_excel('/path/to/data.xlsx', sheet_name='Sheet1', output='records')
    """
    try:
        import pandas as pd
    except ImportError:
        warnings.warn(
            "pandas is not installed so load_excel_light will be used. Data can only be returned as list of dicts."
        )
        return load_excel_light(path, sheets=sheet_name)

    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    data = pd.read_excel(path, sheet_name=sheet_name, dtype=dtype, **kwargs)
    if output == "df":
        return data
    else:
        return {k: v.to_dict(orient=output) for k, v in data.items()}


def write_excel(path, data):
    """
    Write data to an Excel file with multiple sheets.

    Writes a dictionary of data (DataFrames, lists, or dicts) to an Excel file
    with each key becoming a separate sheet. Automatically adjusts column widths.

    :param path: Path where the Excel file should be written
    :type path: str
    :param data: Dictionary where keys are sheet names and values are data to write
    :type data: dict
    :return: None
    :raises FileNotFoundError: If the file path is not an Excel file
    :raises ImportError: If pandas is not installed
    :raises ValueError: If data format is not supported

    Example:
        >>> data = {
        ...     'Sheet1': pd.DataFrame({'A': [1, 2], 'B': [3, 4]}),
        ...     'Sheet2': [{'x': 1, 'y': 2}, {'x': 3, 'y': 4}]
        ... }
        >>> write_excel('/path/to/output.xlsx', data)
    """
    try:
        import pandas as pd
    except ImportError:
        warnings.warn("pandas is not installed so write_excel_light will be used.")
        return write_excel_light(path, data)

    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    with pd.ExcelWriter(path) as writer:
        for sheet_name, content in data.items():
            if isinstance(content, pd.DataFrame):
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif isinstance(content, list):
                df = pd.DataFrame.from_records(content)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif isinstance(content, dict):
                df = pd.DataFrame.from_dict(content, orient="columns")
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            adjust_excel_col_width(writer, df, sheet_name)


def load_csv(path, **kwargs):
    """
    Load a CSV file as a pandas DataFrame.

    Reads a CSV file and returns it as a pandas DataFrame. Falls back to
    the lightweight CSV loader if pandas is not available.

    :param path: Path to the CSV file
    :type path: str
    :param kwargs: Additional keyword arguments passed to pandas.read_csv()
    :return: DataFrame containing the CSV data
    :rtype: pandas.DataFrame
    :raises FileNotFoundError: If the file is not a CSV file
    :raises ImportError: If pandas is not installed

    Example:
        >>> df = load_csv('/path/to/data.csv')
        >>> print(df.head())
    """
    try:
        import pandas as pd
    except ImportError as e:
        warnings.warn("pandas is not installed so load_csv_light will be used.")
        return load_csv_light(path)

    if not check_extension(path, ".csv"):
        raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    return pd.read_csv(path, **kwargs)


def load_csv_light(path, sep=None, encoding=None):
    """
    Load CSV data using the standard csv library (pandas-free).

    Reads a CSV file using Python's built-in csv module and returns
    the data as a list of dictionaries. Automatically detects the
    delimiter if not specified.

    :param path: Path to the CSV file
    :type path: str
    :param sep: Column separator (auto-detected if None)
    :type sep: str, optional
    :param encoding: File encoding (default: system default)
    :type encoding: str, optional
    :return: List of dictionaries representing CSV rows
    :rtype: list[dict]
    :raises ValueError: If the CSV format cannot be determined
    :raises OSError: If the file cannot be read

    Example:
        >>> data = load_csv_light('/path/to/data.csv')
        >>> print(data[0])  # First row as dict
        {'column1': 'value1', 'column2': 'value2'}
    """
    # if not check_extension(path, ".csv"):
    #     raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    with open(path, encoding=encoding) as f:
        try:
            dialect = csv.Sniffer().sniff(f.read(1000), delimiters=sep)
        except Exception as e:
            if sep is not None:
                dialect = get_default_dialect(sep, csv.QUOTE_NONNUMERIC)
            else:
                raise ValueError(f"Error in load_csv_light; {e}")
        dialect.quoting = csv.QUOTE_NONNUMERIC
        f.seek(0)
        if sep is None:
            sep = dialect.delimiter
        else:
            dialect.delimiter = sep
        headers = f.readline().split("\n")[0].split(sep)
        try:
            reader = csv.DictReader(f, dialect=dialect, fieldnames=headers)
            data = [row for row in reader]
        except ValueError:
            warnings.warn("csv loading failed, trying other quote option")
            dialect.quoting = csv.QUOTE_MINIMAL
            reader = csv.DictReader(f, dialect=dialect, fieldnames=headers)
            data = [row for row in reader]

    return data


def get_default_dialect(sep, quoting):
    """
    Create a default CSV dialect with specified separator and quoting.

    Creates a custom CSV dialect with the specified separator and quoting
    style for reading and writing CSV files.

    :param sep: Column separator character
    :type sep: str
    :param quoting: Quoting style (csv.QUOTE_NONNUMERIC, csv.QUOTE_MINIMAL, etc.)
    :type quoting: int
    :return: Configured CSV dialect
    :rtype: csv.Dialect

    Example:
        >>> dialect = get_default_dialect(',', csv.QUOTE_NONNUMERIC)
        >>> reader = csv.DictReader(file, dialect=dialect)
    """

    class dialect(csv.Dialect):
        _name = "default"
        lineterminator = "\r\n"

    dialect.quoting = quoting
    dialect.doublequote = True
    dialect.delimiter = sep
    dialect.quotechar = '"'
    dialect.skipinitialspace = False
    return dialect


def write_csv(path, data, **kwargs):
    """
    Write data to a CSV file.

    Writes data (DataFrame, list of dicts, or dict) to a CSV file.
    Falls back to the lightweight CSV writer if pandas is not available.

    :param path: Path where the CSV file should be written
    :type path: str
    :param data: Data to write (DataFrame, list of dicts, or dict)
    :type data: Union[pandas.DataFrame, list, dict]
    :param kwargs: Additional keyword arguments passed to pandas.to_csv()
    :return: None
    :raises FileNotFoundError: If the file path is not a CSV file
    :raises ImportError: If pandas is not installed

    Example:
        >>> data = [{'name': 'John', 'age': 30}, {'name': 'Jane', 'age': 25}]
        >>> write_csv('/path/to/output.csv', data)
    """
    try:
        import pandas as pd
    except ImportError as e:
        warnings.warn("pandas is not installed so write_csv_light will be used.")
        return write_csv_light(path, data)

    if not check_extension(path, ".csv"):
        raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    if isinstance(data, list):
        df = pd.DataFrame.from_records(data)
        df.to_csv(path, **kwargs, index=False)
    elif isinstance(data, dict):
        df = pd.DataFrame.from_dict(data)
        df.to_csv(path, **kwargs, index=False)


def write_csv_light(path, data, sep=None, encoding=None):
    """
    Write data to CSV using the standard csv library (pandas-free).

    Writes a list of dictionaries to a CSV file using Python's built-in
    csv module. The first dictionary's keys become the column headers.

    :param path: Path where the CSV file should be written
    :type path: str
    :param data: List of dictionaries to write
    :type data: list[dict]
    :param sep: Column separator (default: ',')
    :type sep: str, optional
    :param encoding: File encoding (default: system default)
    :type encoding: str, optional
    :return: None
    :raises FileNotFoundError: If the file path is not a CSV file
    :raises ValueError: If data is empty or invalid

    Example:
        >>> data = [{'name': 'John', 'age': 30}, {'name': 'Jane', 'age': 25}]
        >>> write_csv_light('/path/to/output.csv', data)
    """
    if not check_extension(path, ".csv"):
        raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    if sep is None:
        sep = ","

    dialect = get_default_dialect(sep, csv.QUOTE_NONE)
    with open(path, "w", newline="", encoding=encoding) as f:
        headers = data[0].keys()
        writer = csv.DictWriter(f, fieldnames=headers, dialect=dialect)
        writer.writerow({h: h for h in headers})
        writer.writerows(data)


def adjust_excel_col_width(writer, df, table_name: str, min_len: int = 7):
    """
    Adjust column widths in an Excel file for better readability.

    Automatically adjusts the width of columns in an Excel worksheet
    based on the content length, with a minimum width constraint.

    :param writer: Excel writer object (pandas ExcelWriter)
    :type writer: pandas.ExcelWriter
    :param df: DataFrame containing the data
    :type df: pandas.DataFrame
    :param table_name: Name of the worksheet/sheet
    :type table_name: str
    :param min_len: Minimum column width (default: 7)
    :type min_len: int
    :return: None

    Example:
        >>> with pd.ExcelWriter('output.xlsx') as writer:
        ...     df.to_excel(writer, sheet_name='Sheet1')
        ...     adjust_excel_col_width(writer, df, 'Sheet1')
    """
    for column in df:
        content_len = df[column].astype(str).map(len).max()
        column_width = max(content_len, len(str(column)), min_len) + 2
        col_idx = df.columns.get_loc(column)
        writer.sheets[table_name].set_column(col_idx, col_idx, column_width)


def load_excel_light(path, sheets=None):
    """
    Load an Excel file without pandas dependency.

    Reads an Excel file using openpyxl and returns the data as a dictionary
    of TupLists (list of dictionaries). This is a lightweight alternative
    to the pandas-based Excel loader.

    :param path: Path to the Excel file
    :type path: str
    :param sheets: List of sheet names to read (None for all sheets)
    :type sheets: list, optional
    :return: Dictionary where keys are sheet names and values are TupLists
    :rtype: dict[str, TupList]
    :raises FileNotFoundError: If the file is not an Excel file
    :raises OSError: If the file cannot be read

    Example:
        >>> data = load_excel_light('/path/to/data.xlsx')
        >>> print(data['Sheet1'][0])  # First row of Sheet1
        {'column1': 'value1', 'column2': 'value2'}
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    wb = xl.load_workbook(path, read_only=True, keep_vba=False)

    if sheets is None:
        dict_sheets = {ws.title: ws.values for ws in wb}
    else:
        dict_sheets = {ws.title: ws.values for ws in wb if ws.title in sheets}

    dataset = {}
    for name, v in dict_sheets.items():
        data = [row for row in v]
        if len(data):
            dataset[name] = (
                TupList(data[1:])
                .to_dictlist(data[0])
                .vapply(lambda v: {k: load_str_iterable(w) for k, w in v.items()})
            )
        else:
            dataset[name] = []
    wb.close()
    return dataset


def load_str_iterable(v):
    """
    Parse Excel cell values that represent Python iterables.

    Attempts to evaluate string representations of Python iterables
    (lists, tuples, dicts) in Excel cells and returns them as actual
    Python objects. Other values are returned unchanged.

    :param v: Cell content from Excel
    :type v: Any
    :return: Parsed value (iterable if possible, original value otherwise)
    :rtype: Any

    Example:
        >>> load_str_iterable('[1, 2, 3]')
        [1, 2, 3]
        >>> load_str_iterable('{"key": "value"}')
        {'key': 'value'}
        >>> load_str_iterable('simple string')
        'simple string'
    """
    if isinstance(v, str):
        try:
            return ast.literal_eval(v)
        except (SyntaxError, ValueError):
            # Not a valid Python literal, return as string
            return v
    else:
        return v


def write_excel_light(path, data):
    """
    Write data to an Excel file without pandas dependency.

    Writes a dictionary of data to an Excel file using openpyxl.
    Each key becomes a separate sheet, and the data is formatted
    as tables with automatic column width adjustment.

    :param path: Path where the Excel file should be written
    :type path: str
    :param data: Dictionary where keys are sheet names and values are data
    :type data: dict
    :return: None
    :raises FileNotFoundError: If the file path is not an Excel file
    :raises ValueError: If data format is not supported

    Example:
        >>> data = {
        ...     'Sheet1': [{'A': 1, 'B': 2}, {'A': 3, 'B': 4}],
        ...     'Sheet2': [{'X': 'a', 'Y': 'b'}]
        ... }
        >>> write_excel_light('/path/to/output.xlsx', data)
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    wb = xl.Workbook()
    if len(data):
        wb.remove(wb.active)

    for sheet_name, content in data.items():
        wb.create_sheet(sheet_name)
        if isinstance(content, list):
            ws = wb[sheet_name]
            if len(content):
                ws.append([k for k in content[0].keys()])
                for row in content:
                    ws.append([write_iterables_as_str(v) for v in row.values()])

                tab = get_default_table_style(sheet_name, content)
                ws.add_table(tab)
                adjust_excel_col_width_2(ws)

    wb.save(path)
    wb.close()
    return None


def write_iterables_as_str(v):
    """
    Convert iterables to string representation for Excel cells.

    Converts Python iterables (lists, tuples, dicts) to string
    representation for storage in Excel cells. Non-iterable values
    are returned unchanged.

    :param v: Cell content to convert
    :type v: Any
    :return: String representation if iterable, original value otherwise
    :rtype: Union[str, Any]

    Example:
        >>> write_iterables_as_str([1, 2, 3])
        '[1, 2, 3]'
        >>> write_iterables_as_str('simple string')
        'simple string'
    """
    if isinstance(v, Iterable):
        return str(v)
    else:
        return v


def get_default_table_style(sheet_name, content):
    """
    Create a default table style for Excel worksheets.

    Generates a default table style configuration for Excel worksheets
    with basic formatting options.

    :param sheet_name: Name of the worksheet
    :type sheet_name: str
    :param content: List of dictionaries representing the table data
    :type content: list[dict]
    :return: Configured table object
    :rtype: openpyxl.worksheet.table.Table

    Example:
        >>> content = [{'A': 1, 'B': 2}, {'A': 3, 'B': 4}]
        >>> table = get_default_table_style('Sheet1', content)
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    style = TableStyleInfo(
        name="style_1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    tab = Table(
        displayName=sheet_name,
        ref=f"A1:{get_column_letter(len(content[0]))}{len(content) + 1}",
    )
    tab.tableStyleInfo = style
    return tab


def adjust_excel_col_width_2(ws, min_width=10, max_width=30):
    """
    Adjust column widths in an Excel worksheet with constraints.

    Automatically adjusts column widths based on content length with
    minimum and maximum width constraints for better readability.

    :param ws: Excel worksheet object
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :param min_width: Minimum column width (default: 10)
    :type min_width: int
    :param max_width: Maximum column width (default: 30)
    :type max_width: int
    :return: None

    Example:
        >>> ws = wb['Sheet1']
        >>> adjust_excel_col_width_2(ws, min_width=8, max_width=25)
    """
    for k, v in get_column_widths(ws).items():
        ws.column_dimensions[k].width = min(max(v, min_width), max_width)
    return None


def get_column_widths(ws):
    """
    Calculate optimal column widths for an Excel worksheet.

    Analyzes the content of each column in a worksheet and returns
    the optimal width for each column based on the longest content.

    :param ws: Excel worksheet object
    :type ws: openpyxl.worksheet.worksheet.Worksheet
    :return: Dictionary mapping column letters to their optimal widths
    :rtype: dict[str, float]

    Example:
        >>> ws = wb['Sheet1']
        >>> widths = get_column_widths(ws)
        >>> print(widths)
        {'A': 15.0, 'B': 12.0, 'C': 20.0}
    """
    result = {}
    for column_cells in ws.columns:
        width = max(len(str(cell.value)) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        result[letter] = width * 1.2
    return result
