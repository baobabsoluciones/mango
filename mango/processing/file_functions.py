import json
from os import listdir
from typing import Union

import pandas as pd
import openpyxl as xl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from pytups import TupList
from openpyxl.utils import get_column_letter


def list_files_directory(directory: str, extensions: list = None):
    """
    The list_files_directory function returns a list of files in the directory specified by the user.
    The function takes two arguments:
        1) The directory to search for files in (str).
        2) A list of file extensions to filter by (list). If no extensions are provided, all files will be returned.

    :param str directory: Specify the directory that you want to list files from
    :param list extensions: Specify the file extensions that should be included
    :return: A list of all filtered files in a directory
    :raises WindowsError: if the directory doesn't exist
    """
    if extensions is None:
        extensions = ["."]
    return [
        rf"{directory}\{f}"
        for f in listdir(rf"{directory}")
        if any([f.__contains__(f"{ext}") for ext in extensions])
    ]


def check_extension(path: str, extension: str):
    """
    The check_extension function checks if a file has the specified extension.

    :param path: Specify the path of the file to be checked
    :param extension: Specify the extension to check against
    :return: A boolean
    :doc-author: baobab soluciones
    """
    return path.endswith(extension)


def is_excel_file(path: str):
    """
    The is_excel_file function checks if a file is an Excel file.

    :param path: Specify the path of the file to be checked
    :return: A boolean
    :doc-author: baobab soluciones
    """
    return (
        check_extension(path, ".xlsx")
        or check_extension(path, ".xls")
        or check_extension(path, ".xlsm")
    )


def is_json_file(path: str):
    """
    The is_json_file function checks if a file is a JSON file.

    :param path: Specify the path of the file to be checked
    :return: A boolean
    :doc-author: baobab soluciones
    """
    return check_extension(path, ".json")


def load_json(path: str, **kwargs):
    """
    The load_json function loads a json file from the specified path and returns it as a dictionary.

    :param path: Specify the path of the file to be loaded
    :return: A dictionary
    :doc-author: baobab soluciones
    """
    with open(path, "r", **kwargs) as f:
        return json.load(f)


def write_json(data: Union[dict, list], path):
    """
    The write_json function writes a dictionary or list to a JSON file.

    :param data: allow the function to accept both a dictionary and list object
    :param path: Specify the path of the file that you want to write to
    :return: None
    :doc-author: baobab soluciones
    """

    with open(path, "w") as f:
        json.dump(data, f, indent=4, sort_keys=False)


def load_excel_sheet(path: str, sheet: str, **kwargs):
    """
    The load_excel_sheet function loads a sheet from an Excel file and returns it as a DataFrame.

    :param path: Specify the path of the file to be loaded
    :param sheet: Specify the name of the sheet to be loaded
    :return: A DataFrame
    :doc-author: baobab soluciones
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    return pd.read_excel(path, sheet_name=sheet, **kwargs)


def load_excel(path):
    """
    The load_excel function loads an Excel file and returns it as a dictionary of DataFrames.

    :param path: Specify the path of the file to be loaded
    :return: A dictionary of DataFrames
    :doc-author: baobab soluciones
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )
    return pd.read_excel(path, sheet_name=None)


def write_excel(path, data):
    """
    The write_excel function writes a dictionary of DataFrames to an Excel file.

    :param path: Specify the path of the file that you want to write to
    :param data: Specify the dictionary to be written
    :return: None
    :doc-author: baobab soluciones
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    with pd.ExcelWriter(path) as writer:
        for sheet_name, content in data.items():
            if isinstance(content, list):
                df = pd.DataFrame.from_records(content)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            elif isinstance(content, dict):
                df = pd.DataFrame.from_dict(content, orient="index")
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            adjust_excel_col_width(writer, df, sheet_name)


def write_df_to_excel(path, data, **kwargs):
    """
    The write_df_to_excel function writes a DataFrame to an Excel file.

    :param path: Specify the path of the file that you want to write to
    :param data: Specify the DataFrame to be written
    :return: None
    :doc-author: baobab soluciones
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    with pd.ExcelWriter(path) as writer:
        data.to_excel(writer, **kwargs, index=False)


def load_csv(path, **kwargs):
    """
    The load_csv function loads a CSV file and returns it as a DataFrame.

    :param path: Specify the path of the file to be loaded
    :return: A DataFrame
    :doc-author: baobab soluciones
    """
    if not check_extension(path, ".csv"):
        raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    return pd.read_csv(path, **kwargs)


def write_csv(path, data, **kwargs):
    """
    The write_csv function writes a DataFrame to a CSV file.

    :param path: Specify the path of the file that you want to write to
    :param data: Specify the DataFrame to be written
    :return: None
    :doc-author: baobab soluciones
    """
    if not check_extension(path, ".csv"):
        raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    if isinstance(data, list):
        df = pd.DataFrame.from_records(data)
        df.to_csv(path, **kwargs, index=False)
    elif isinstance(data, dict):
        df = pd.DataFrame.from_dict(data)
        df.to_csv(path, **kwargs, index=False)


def write_df_to_csv(path, data, **kwargs):
    """
    The write_csv function writes a DataFrame to a CSV file.

    :param path: Specify the path of the file that you want to write to
    :param data: Specify the DataFrame to be written
    :return: None
    :doc-author: baobab soluciones
    """
    if not check_extension(path, ".csv"):
        raise FileNotFoundError(f"File {path} is not a CSV file (.csv).")

    data.to_csv(path, **kwargs, index=False)


def adjust_excel_col_width(writer, df: pd.DataFrame, table_name: str, min_len: int = 7):
    """
    Adjusts the width of the column on the Excel file

    :param writer:
    :param :class:`pandas.DataFrame` df:
    :param str table_name:
    :param int min_len:
    """
    for column in df:
        content_len = df[column].astype(str).map(len).max()
        column_width = max(content_len, len(str(column)), min_len) + 2
        col_idx = df.columns.get_loc(column)
        writer.sheets[table_name].set_column(col_idx, col_idx, column_width)


def load_excel_2(path):
    """
    The load_excel function loads an Excel file and returns it as a dictionary of DataFrames.
    It doesn't use pandas.

    :param path: Specify the path of the file to be loaded
    :return: A dictionary of TupLists
    :doc-author: baobab soluciones
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )

    wb = xl.load_workbook(path)
    dataset = {}
    for ws in wb:
        data = [row for row in ws.values]
        if len(data):
            dataset[ws.title] = TupList(data[1:]).to_dictlist(data[0])
        else:
            dataset[ws.title] = []

    return dataset


def write_excel_2(path, data):
    """
    The write_excel function writes a dictionary of DataFrames to an Excel file.

    :param path: Specify the path of the file that you want to write to
    :param data: Specify the dictionary to be written
    :return: None
    :doc-author: baobab soluciones
    """
    if not is_excel_file(path):
        raise FileNotFoundError(
            f"File {path} is not an Excel file (.xlsx, .xls, .xlsm)."
        )
    wb = xl.Workbook()
    if len(data):
        wb.remove(wb.active)

    for sheet_name, content in data.items():
        wb.create_sheet(sheet_name)
        if isinstance(content, list):
            ws = wb[sheet_name]
            if len(content):
                ws.append([k for k in content[0].keys()])
                for row in content:
                    ws.append([v for v in row.values()])

                tab = get_default_table_style(sheet_name, content)

                ws.add_table(tab)
                adjust_excel_col_width_2(ws)

    wb.save(path)
    return None


def get_default_table_style(sheet_name, content):
    """
    Get a default style for the table

    :param sheet_name: name of the sheet
    :param content: content of the sheet.
    :return: table object
    """
    from openpyxl.worksheet.table import Table, TableStyleInfo

    #  TODO: look at interesting styles
    style = TableStyleInfo(
        name="style_1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    tab = Table(
        displayName=sheet_name,
        ref=f"A1:{get_column_letter(len(content[0]))}{len(content) + 1}",
    )
    tab.tableStyleInfo = style
    return tab


def adjust_excel_col_width_2(ws, min_width=10, max_width=30):
    """
    Adjust the column width of a worksheet.

    :param ws: worksheet object
    :param min_width: minimum width to use.
    :param max_width: maximum width to use
    :return: None
    """
    for k, v in get_column_widths(ws).items():
        ws.column_dimensions[k].width = min(max(v, min_width), max_width)
    return None


def get_column_widths(ws):
    """
    Get the maximum width of the columns of a worksheet.

    :param ws: worksheet object
    :return: a dict with the letter of the columns and their maximum width (ex: {"A":15, "B":12})
    """
    result = {}
    for column_cells in ws.columns:
        width = max(len(str(cell.value)) for cell in column_cells)
        letter = get_column_letter(column_cells[0].column)
        result[letter] = width * 1.2
    return result


# data = load_excel_2("./test.xlsx")
#
# write_excel_2("./test2.xlsx", data)
